import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from io import BytesIO

# --- Funciones auxiliares (igual que antes) ---

def validar_rango(valor, minimo=0, maximo=1):
    try:
        val = float(valor)
        return max(min(val, maximo), minimo)
    except:
        return minimo

def clasificar_amenaza(val):
    val = float(val)
    if val <= 0.05:
        return "Rara"
    elif val <= 0.15:
        return "Poco Probable"
    elif val <= 0.4:
        return "Posible"
    elif val <= 0.7:
        return "Probable"
    else:
        return "Casi Seguro"

def clasificar_impacto(val):
    val = float(val)
    if val <= 5:
        return "Insignificante"
    elif val <= 15:
        return "Leve"
    elif val <= 30:
        return "Moderado"
    elif val <= 60:
        return "Grave"
    else:
        return "CrÃ­tico"

def clasificar_aceptabilidad(val):
    val = float(val)
    if val <= 0.7:
        return "Aceptable"
    elif val <= 3:
        return "Tolerable"
    elif val <= 7:
        return "Inaceptable"
    else:
        return "Inadmisible"

def evaluar_efectividad(efectividad):
    ef = float(efectividad)
    if ef == 0:
        return "Inefectiva"
    elif ef <= 0.2:
        return "Limitada"
    elif ef <= 0.4:
        return "Baja"
    elif ef <= 0.6:
        return "Intermedia"
    elif ef <= 0.81:
        return "Alta"
    elif ef <= 0.95:
        return "Muy alta"
    else:
        return "Total"

def aplicar_color_criticidad(valor):
    val = float(valor)
    if val <= 2:
        return "00FF00"  # Verde
    elif val <= 4:
        return "FFFF00"  # Amarillo
    elif val <= 15:
        return "FFA500"  # Naranja
    else:
        return "FF0000"  # Rojo

# --- FunciÃ³n que procesa el workbook en memoria ---

def procesar_workbook(wb):
    ws = wb.active
    max_fila = ws.max_row

    for bloque_inicio in range(23, max_fila + 1, 14):
        bloque_fin = min(bloque_inicio + 8, max_fila)
        suma_W = 0

        for fila in range(bloque_inicio, bloque_fin + 1):
            efectividad = validar_rango(ws[f"I{fila}"].value)
            exposicion = validar_rango(ws[f"L{fila}"].value)
            probabilidad = validar_rango(ws[f"N{fila}"].value)
            amenaza_inherente = validar_rango(ws[f"O{fila}"].value or 0)

            Q = amenaza_inherente * exposicion * probabilidad * (1 - efectividad)
            S = Q * (1 + 0.25 * (1 - efectividad))
            impacto = float(ws[f"U{fila}"].value or 0.05)
            W = S * impacto

            ws[f"Q{fila}"].value = Q
            ws[f"R{fila}"].value = clasificar_amenaza(Q)
            ws[f"S{fila}"].value = S
            ws[f"T{fila}"].value = clasificar_amenaza(S)
            ws[f"W{fila}"].value = W
            ws[f"V{fila}"].value = clasificar_impacto(impacto)
            ws[f"Y{fila}"].value = clasificar_aceptabilidad(W)
            ws[f"K{fila}"].value = evaluar_efectividad(efectividad)

            suma_W += W

        ws[f"Z{bloque_inicio}"].value = suma_W
        ws[f"AA{bloque_inicio}"].value = suma_W / 294 * 100

        factor_deliberada = 1.4
        valor_ac_celda = ws[f"AC{bloque_inicio + 8}"].value
        if valor_ac_celda and isinstance(valor_ac_celda, (int, float)) and 1 <= valor_ac_celda <= 3:
            factor_deliberada = valor_ac_celda

        valor_ac = ws[f"AA{bloque_inicio}"].value * factor_deliberada
        ws[f"AC{bloque_inicio}"].value = valor_ac

        color = aplicar_color_criticidad(valor_ac)
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws[f"AC{bloque_inicio}"].fill = fill
        ws[f"AC{bloque_inicio}"].font = Font(bold=True)
        ws[f"AA{bloque_inicio}"].font = Font(bold=True)

    return wb

# --- Streamlit UI ---

st.title("Calculadora de riesgos - EvaluaciÃ³n desde Excel")

uploaded_file = st.file_uploader("Sube tu archivo Excel (.xlsm)", type=["xlsm", "xlsx"])

if uploaded_file is not None:
    try:
        wb = load_workbook(filename=BytesIO(uploaded_file.read()), keep_vba=True)
        wb = procesar_workbook(wb)

        # Guardar resultado en memoria para descarga
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        st.success("Archivo procesado correctamente ðŸŽ‰")

        st.download_button(
            label="Descargar archivo procesado",
            data=output,
            file_name="resultado_evaluacion_riesgos.xlsm",
            mime="application/vnd.ms-excel"
        )
    except Exception as e:
        st.error(f"Error procesando archivo: {e}")

